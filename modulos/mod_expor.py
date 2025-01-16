import openpyxl
import sqlite3
from modulos.mod_base import interc
from datetime import datetime
#realizar consulta de los datos de la tabla gastos e importarlos a excel con openpyxl
#conectar a la base de datos

#inicializar
i = interc()
def exportar(nom):
    #nom es el nombre de la tabla a consultar
    datos = i.tabla(nom)
    # Crear un libro de Excel
    wb = openpyxl.Workbook()
    #obtener la hoja activa
    ws = wb.active
    #asignar el título
    ws.title = nom
    #   asignar los títulos de las columnas
    ws['A1'] = 'ID'
    ws['B1'] = "Concepto"
    ws['C1'] = "Monto"
    ws['D1'] = "Fecha Gasto"
    ws['E1'] = "Fecha Registro"
    ws['F1'] = "Persona que realizo"
    ws['G1'] = "Quien recibe"
    #insertar los datos
    for b in range(len(datos)):
        ws.append(datos[b])
    #darle tamaño a las celdas correspondiente
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15  
    #guardar el archivo
    
    #apartado opcional solo aplica en gastos
    #se puede agregar formato a la celda
    print("Requiere total? 1. si es asi ponga la celda que sera 2.No")
    op_t = int(input())
    if op_t == 1:
        total = 0
        columna = int(input("Ingrese #de columna"))
        for d in range(len(datos)):
            total += int(datos[d][2])
        #agregarlo despues del ultimo registro de la columna
        ws.cell(row=len(datos)+2, column=4).value = total
    
    wb.save(f"{nom}_{datetime.now().strftime("%Y%m%d")}.xlsx")

def menu_exportar():
    print("Que desea exportar")
    tablas = i.consulta_tablas()
    for x, tabla in enumerate(tablas,start=0):
        print(f"{x+1}. {tabla}")
    tabla = int(input())
    exportar(tablas[tabla-1])
    #exportar("gastos")